#!/usr/bin/env python3
"""식약처 식품영양성분DB 엑셀을 `build-food-csv.py`가 읽을 CSV로 바꾼다.

공공데이터포털이 주는 원본이 xlsx라서 한 단계가 더 필요하다. 160개가 넘는 컬럼 중
정제에 쓰는 것만 남겨 중간 파일이 커지지 않게 한다(가공식품 원본은 188MB다).

  python3 scripts/xlsx-to-csv.py ~/Downloads/음식DB.xlsx /tmp/food-raw.csv
  python3 scripts/build-food-csv.py /tmp/food-raw.csv \
    apps/daily-record/src/main/resources/food/food-nutrition.csv

`.xlsx`는 `openpyxl`, `.xls`는 `xlrd`가 필요하다: `pip3 install openpyxl xlrd`
(포털이 데이터셋마다 다른 형식을 준다 — 원재료성식품은 `.xls`다)
"""

import csv
import sys

import openpyxl

# 남길 컬럼. 헤더 표기가 판본마다 조금씩 달라 공백을 지운 뒤 부분 문자열로 찾는다.
# 순서가 곧 출력 순서이며, 앞에 있는 힌트가 먼저 매칭된다.
WANTED = [
    "식품코드",
    "식품명",
    "영양성분함량기준량",
    "에너지(kcal)",
    "탄수화물(g)",
    "단백질(g)",
    "지방(g)",
    "당류(g)",
    "나트륨(mg)",
    "식이섬유(g)",
    # **`지방(g)` 뒤여야 한다.** 이 순서가 곧 중간 CSV의 컬럼 순서이고,
    # `build-food-csv.py`의 `fat` 힌트("지방")는 중간 CSV를 앞에서부터 훑는다.
    # 앞에 두면 포화지방산이 지방으로 읽혀 지방 값이 통째로 틀어진다 — 오류는 나지 않는다.
    "포화지방산(g)",
    "트랜스지방산(g)",
    "콜레스테롤(mg)",
    "1인(회)분량참고량",
    "식품중량",
    # **프랜차이즈 브랜드가 여기 있다.** 음식DB는 `업체명`, 가공식품DB는 `제조사명`이고
    # 데이터셋마다 하나만 존재한다. 둘 다 넣어 두고 `build-food-csv.py`가 있는 쪽을 고른다.
    # 이 컬럼이 없으면 도미노피자 318건이 적재돼 있어도 「도미노」로 찾을 수 없다 —
    # 식품명이 `피자_뉴욕 오리진 피자 오리지널 (L)`이라 브랜드가 어디에도 없기 때문이다.
    "업체명",
    "제조사명",
    # 원재료성식품에만 있다. `식품명`이 `복숭아_천중도_생것`처럼 계층형이라 완전일치가 안 되므로
    # 대표명으로 묶고, 같은 대표명 안에서 `생것`만 남긴다(`build-food-csv.py --representative`).
    "대표식품명",
    "식품세분류명",
]

# 없어도 되는 컬럼 — 데이터셋에 따라 빠져 있다. `build-food-csv.py`의 OPTIONAL_COLUMNS와 정책을
# 맞춘다 — 당류·나트륨·식이섬유는 판본이 바뀌어 컬럼이 빠지더라도 탄단지는 살려야 한다.
OPTIONAL = {
    "1인(회)분량참고량",
    "식품중량",
    "당류(g)",
    "나트륨(mg)",
    "식이섬유(g)",
    "포화지방산(g)",
    "트랜스지방산(g)",
    "콜레스테롤(mg)",
    "업체명",
    "제조사명",
    "대표식품명",
    "식품세분류명",
}


def find_column(header, hint):
    """**완전일치를 먼저 찾고**, 없으면 원본 컬럼 순서대로 부분 문자열로 찾는다.

    부분 문자열만 쓰면 포함 관계인 헤더에서 엉뚱한 컬럼이 잡힌다 — `업체명`은 가공식품DB의
    `수입업체명`·`유통업체명`에도 들어 있고, `포화지방산(g)`은 `불포화지방산(g)`에도 들어 있다.
    완전일치를 우선하면 그런 경우가 대부분 걸러지고, 판본에 따라 표기가 조금 다를 때만
    부분일치로 떨어진다.

    부분일치 순서는 여전히 **원본 컬럼 순서**다 — `지방(g)` 힌트가 뒤쪽 파생 컬럼보다
    앞의 본 컬럼을 잡게 하기 위해서다.
    """
    squeezed = hint.replace(" ", "")
    for index, name in enumerate(header):
        if squeezed == name.replace(" ", ""):
            return index
    for index, name in enumerate(header):
        if squeezed in name.replace(" ", ""):
            return index
    return None


def read_rows(src_path):
    """행을 튜플로 흘려보낸다. `.xls`(구형 OLE2)와 `.xlsx`(OOXML)는 읽는 라이브러리가 다르다.

    포털이 데이터셋마다 다른 형식을 준다 — 음식·가공식품은 `.xlsx`인데 원재료성식품은 `.xls`다.
    `openpyxl`은 `.xls`를 못 읽고 오류 메시지도 형식 문제임을 알려주지 않아서, 확장자로 갈라
    `xlrd`를 쓴다(`pip3 install xlrd`).
    """
    if src_path.lower().endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(src_path)
        sheet = book.sheet_by_index(0)
        for index in range(sheet.nrows):
            yield tuple(cell.value for cell in sheet.row(index))
        return

    workbook = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    yield from workbook[workbook.sheetnames[0]].iter_rows(values_only=True)


def main(src_path, dst_path):
    rows = read_rows(src_path)

    # 헤더가 첫 줄이 아닌 판본이 있다 — 원재료성식품 `.xls`는 1행이 통째로 비어 있고 2행이 헤더다.
    header = []
    while not any(header):
        header = ["" if v is None else str(v).strip() for v in next(rows)]
    found = [(hint, find_column(header, hint)) for hint in WANTED]
    missing = [hint for hint, index in found if index is None and hint not in OPTIONAL]
    if missing:
        sys.exit(f"원본에서 컬럼을 찾지 못했습니다: {missing}\n헤더: {header[:20]} ...")

    # **못 찾은 선택 컬럼은 아예 내보내지 않는다.** 빈 자리표시자를 쓰면 `build-food-csv.py`가
    # 「있는데 값이 전부 빈 컬럼」으로 보고 골라 버린다 — 음식DB에 없는 `제조사명`이 자리표시자로
    # 들어가 `업체명`보다 먼저 잡히면서 브랜드가 통째로 비었던 자리다.
    columns = [(hint, index) for hint, index in found if index is not None]
    skipped = [hint for hint, index in found if index is None]

    written = 0
    with open(dst_path, "w", encoding="utf-8", newline="") as dst:
        writer = csv.writer(dst)
        writer.writerow([header[index] for _, index in columns])
        for row in rows:
            # 식품코드가 없는 줄은 빈 줄이거나 꼬리말이다.
            code_index = columns[0][1]
            if row[code_index] is None or not str(row[code_index]).strip():
                continue
            writer.writerow(
                ["" if row[index] is None else str(row[index]).strip() for _, index in columns]
            )
            written += 1

    if skipped:
        print(f"이 원본에 없어 건너뛴 선택 컬럼: {skipped}")

    print(f"{written}행을 {dst_path}에 썼습니다 (원본 컬럼 {len(header)}개 중 {len(columns)}개 유지)")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("usage: xlsx-to-csv.py <원본.xlsx> <출력.csv>")
    main(sys.argv[1], sys.argv[2])
